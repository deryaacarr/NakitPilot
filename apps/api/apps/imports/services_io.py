"""File I/O helpers for invoice imports (NP-061 / NP-152)."""

from __future__ import annotations

import csv
import hashlib
import io
import re
import uuid
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from django.conf import settings
from openpyxl import Workbook, load_workbook

from apps.imports.schema import (
    CANONICAL_COLUMNS,
    normalize_header,
)

MAX_UPLOAD_BYTES = 10 * 1024 * 1024  # 10 MB
ALLOWED_EXTENSIONS = {".xlsx", ".xls", ".csv"}
ALLOWED_CONTENT_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
    "text/csv",
    "application/csv",
    "text/plain",
}
# Map sniffed kind → allowed extensions / MIME families
SNIFFED_KINDS = {
    "xlsx": {
        "exts": {".xlsx"},
        "mimes": {
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        },
    },
    "xls": {
        "exts": {".xls"},
        "mimes": {"application/vnd.ms-excel"},
    },
    "csv": {
        "exts": {".csv"},
        "mimes": {"text/csv", "application/csv", "text/plain"},
    },
}
DANGEROUS_NAME_RE = re.compile(r"(\.\.|[\x00-\x1f]|[<>:\"|?*]|/|\\)")


class UploadValidationError(Exception):
    def __init__(self, message: str, code: str = "invalid_file"):
        super().__init__(message)
        self.message = message
        self.code = code


def build_invoice_template_bytes() -> bytes:
    """NP-060: sample Excel workbook with canonical headers + example row."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Faturalar"
    ws.append(list(CANONICAL_COLUMNS))
    ws.append(
        [
            "C-001",
            "Örnek Ticaret A.Ş.",
            "",
            "FTR-2026-001",
            "2026-07-01",
            "2026-07-31",
            "TRY",
            "1500.00",
            "0.00",
            "05551234567",
            "ornek@firma.com",
        ]
    )
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def sanitize_filename(name: str) -> str:
    """Validate original client filename (never used as storage path)."""
    name = (name or "").strip()
    if not name:
        raise UploadValidationError("Dosya adı boş olamaz.", "dangerous_filename")
    if DANGEROUS_NAME_RE.search(name) or name.startswith("."):
        raise UploadValidationError("Zararlı dosya adı.", "dangerous_filename")
    name = Path(name).name
    if DANGEROUS_NAME_RE.search(name):
        raise UploadValidationError("Zararlı dosya adı.", "dangerous_filename")
    if len(name) > 200:
        raise UploadValidationError("Dosya adı çok uzun.", "dangerous_filename")
    return name


def sniff_file_kind(content: bytes) -> str | None:
    """Detect real type from magic bytes / content — do not trust extension alone."""
    if content.startswith(b"PK\x03\x04") or content.startswith(b"PK\x05\x06"):
        return "xlsx"
    if content.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"):
        return "xls"
    # CSV / plain text heuristic
    sample = content[:4096]
    if b"\x00" in sample:
        return None
    try:
        text = sample.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = sample.decode("iso-8859-9")
        except UnicodeDecodeError:
            return None
    if not text.strip():
        return None
    # Reject if it looks like HTML/script
    lower = text.lstrip().lower()
    if lower.startswith("<!doctype") or lower.startswith("<html") or lower.startswith("<?php"):
        return None
    if "," in text or ";" in text or "\t" in text or "\n" in text:
        return "csv"
    return "csv" if text.isprintable() or "\n" in text else None


def validate_upload_file(*, filename: str, size: int, content_type: str, content: bytes) -> str:
    """
    NP-152: extension + declared MIME + sniffed content must agree.
    Returns the sanitized *original* filename for display only.
    """
    filename = sanitize_filename(filename)
    ext = Path(filename).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise UploadValidationError(
            "Desteklenen dosya türleri: .xlsx, .xls, .csv",
            "invalid_file_type",
        )
    if size <= 0 or len(content) == 0:
        raise UploadValidationError("Dosya boş olamaz.", "empty_file")
    if size > MAX_UPLOAD_BYTES or len(content) > MAX_UPLOAD_BYTES:
        raise UploadValidationError("Dosya boyutu en fazla 10 MB olabilir.", "file_too_large")

    ct = (content_type or "").split(";")[0].strip().lower()
    if ct and ct not in ALLOWED_CONTENT_TYPES:
        raise UploadValidationError("Geçersiz içerik türü.", "invalid_file_type")

    kind = sniff_file_kind(content)
    if kind is None:
        raise UploadValidationError("Dosya içeriği tanınamadı.", "invalid_file_type")
    rules = SNIFFED_KINDS[kind]
    if ext not in rules["exts"]:
        raise UploadValidationError(
            "Dosya uzantısı içerikle uyuşmuyor.",
            "invalid_file_type",
        )
    if ct and ct not in rules["mimes"]:
        raise UploadValidationError(
            "MIME türü dosya içeriğiyle uyuşmuyor.",
            "invalid_file_type",
        )
    return filename


def file_sha256(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def private_upload_root() -> Path:
    """
    NP-152: store uploads outside web-served MEDIA/STATIC trees.
    Override with PRIVATE_UPLOAD_ROOT env / setting.
    """
    configured = getattr(settings, "PRIVATE_UPLOAD_ROOT", None)
    if configured:
        return Path(configured)
    return Path(settings.BASE_DIR) / "private_uploads"


def store_upload(*, organization_id: int, filename: str, content: bytes) -> str:
    """
    Persist upload under private root with a regenerated opaque name.
    Never use the client-supplied path.
    """
    ext = Path(sanitize_filename(filename)).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        ext = ".bin"
    root = private_upload_root() / "org" / str(organization_id) / "imports"
    root.mkdir(parents=True, exist_ok=True)
    stored_name = f"{uuid.uuid4().hex}{ext}"
    path = root / stored_name
    # Refuse path escape even if somehow misconfigured
    path = path.resolve()
    if not str(path).startswith(str(root.resolve())):
        raise UploadValidationError("Geçersiz depolama yolu.", "invalid_file")
    path.write_bytes(content)
    return str(path)


def read_tabular_file(path: str | Path) -> tuple[list[str], list[dict[str, Any]]]:
    path = Path(path)
    # Only read from private upload root (or legacy MEDIA imports) when configured
    resolved = path.resolve()
    allowed_roots = [private_upload_root().resolve()]
    media_imports = Path(settings.MEDIA_ROOT).resolve() / "imports"
    allowed_roots.append(media_imports)
    if not any(str(resolved).startswith(str(root)) for root in allowed_roots):
        raise UploadValidationError("Dosya yolu geçersiz.", "invalid_file")
    ext = path.suffix.lower()
    if ext == ".csv":
        return _read_csv(path)
    if ext in {".xlsx", ".xls"}:
        return _read_xlsx(path)
    raise UploadValidationError("Desteklenmeyen dosya türü.", "invalid_file_type")


def _read_xlsx(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration as exc:
        raise UploadValidationError("Dosya boş olamaz.", "empty_file") from exc

    headers = [_cell_to_str(cell) for cell in header_row]
    normalized_headers: list[str] = []
    for idx, header in enumerate(headers):
        label = header.strip() if header else f"kolon_{idx + 1}"
        normalized_headers.append(label)

    if not any(h.strip() for h in normalized_headers):
        raise UploadValidationError("Başlık satırı bulunamadı.", "missing_headers")

    data: list[dict[str, Any]] = []
    for row in rows_iter:
        if row is None or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        record: dict[str, Any] = {}
        for idx, header in enumerate(normalized_headers):
            value = row[idx] if idx < len(row) else None
            record[header] = value
        data.append(record)
    wb.close()
    if not data:
        raise UploadValidationError("Dosyada veri satırı yok.", "empty_file")
    return normalized_headers, data


def _read_csv(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "iso-8859-9", "cp1254"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            text = None
    else:
        raise UploadValidationError("CSV kodlaması okunamadı.", "invalid_file")

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel

    reader = csv.reader(io.StringIO(text), dialect)
    try:
        header_row = next(reader)
    except StopIteration as exc:
        raise UploadValidationError("Dosya boş olamaz.", "empty_file") from exc

    headers = [
        (h or f"kolon_{i + 1}").strip() or f"kolon_{i + 1}" for i, h in enumerate(header_row)
    ]
    data: list[dict[str, Any]] = []
    for row in reader:
        if not row or all(not str(cell).strip() for cell in row):
            continue
        record = {headers[i]: (row[i] if i < len(row) else "") for i in range(len(headers))}
        data.append(record)
    if not data:
        raise UploadValidationError("Dosyada veri satırı yok.", "empty_file")
    return headers, data


def _cell_to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def ensure_required_headers_present(headers: list[str]) -> None:
    """
    NP-061: if file looks like the official template, require canonical columns.
    Otherwise allow custom headers (mapping screen).
    """
    normalized = {normalize_header(h) for h in headers}
    template_hits = sum(
        1 for col in CANONICAL_COLUMNS if normalize_header(col) in normalized or col in headers
    )
    if template_hits >= 5:
        missing = [col for col in CANONICAL_COLUMNS if col not in headers]
        if missing:
            raise UploadValidationError(
                f"Zorunlu kolonlar eksik: {', '.join(missing)}",
                "missing_required_columns",
            )


def parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    return None


def parse_money(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    text = str(value).strip().replace(" ", "").replace("₺", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def mapped_row(row: dict[str, Any], mapping: dict[str, str | None]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for canonical, source in mapping.items():
        if not source:
            result[canonical] = None
            continue
        result[canonical] = row.get(source)
    return result
