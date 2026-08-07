import io
from datetime import date
from decimal import Decimal

import pytest
from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import Workbook, load_workbook
from rest_framework import status
from rest_framework.test import APIClient

from apps.customers.models import Customer
from apps.imports.models import DuplicatePolicy, ImportJobStatus
from apps.imports.schema import CANONICAL_COLUMNS, suggest_mapping
from apps.imports.services import build_invoice_template_bytes, sanitize_filename
from apps.imports.validation import validate_mapped_row
from apps.invoices.models import Invoice, InvoiceStatus
from apps.organizations.models import Membership, Organization, Role

User = get_user_model()
PASSWORD = "SecretPass123!"


@pytest.fixture
def api_client():
    return APIClient()


def _auth(client, user, organization):
    login = client.post(
        "/api/auth/login",
        {"email": user.email, "password": PASSWORD},
        format="json",
    )
    client.credentials(HTTP_AUTHORIZATION=f"Bearer {login.data['access']}")
    client.defaults["HTTP_X_ORGANIZATION_ID"] = str(organization.id)
    return client


@pytest.fixture
def org_owner(db):
    org = Organization.objects.create(name="Import Co", slug="import-co")
    owner = User.objects.create_user(email="import@example.com", password=PASSWORD)
    Membership.objects.create(organization=org, user=owner, role=Role.OWNER, is_active=True)
    return org, owner


def _xlsx_bytes(headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _full_mapping(headers_map: dict[str, str]):
    mapping = {col: None for col in CANONICAL_COLUMNS}
    mapping.update(headers_map)
    return mapping


def test_template_contains_canonical_columns():
    content = build_invoice_template_bytes()
    assert content[:2] == b"PK"  # zip/xlsx


def test_sanitize_rejects_path_traversal():
    with pytest.raises(Exception) as exc:
        sanitize_filename("../secret.xlsx")
    assert exc.value.code == "dangerous_filename"


@pytest.mark.django_db
def test_download_template(api_client, org_owner):
    org, owner = org_owner
    client = _auth(api_client, owner, org)
    response = client.get("/api/imports/invoices/template/")
    assert response.status_code == status.HTTP_200_OK
    assert "spreadsheetml" in response["Content-Type"]
    assert response.content[:2] == b"PK"


@pytest.mark.django_db
def test_upload_validations(api_client, org_owner):
    org, owner = org_owner
    client = _auth(api_client, owner, org)

    empty = client.post(
        "/api/imports/invoices/upload/",
        {"file": SimpleUploadedFile("empty.xlsx", b"", content_type="application/octet-stream")},
        format="multipart",
    )
    assert empty.status_code == status.HTTP_400_BAD_REQUEST
    assert empty.data["code"] == "empty_file"

    bad_type = client.post(
        "/api/imports/invoices/upload/",
        {
            "file": SimpleUploadedFile(
                "evil.exe",
                b"MZ1234",
                content_type="application/octet-stream",
            )
        },
        format="multipart",
    )
    assert bad_type.status_code == status.HTTP_400_BAD_REQUEST
    assert bad_type.data["code"] == "invalid_file_type"

    bad_name = client.post(
        "/api/imports/invoices/upload/",
        {
            "file": SimpleUploadedFile(
                "secret<>.xlsx",
                _xlsx_bytes(
                    list(CANONICAL_COLUMNS),
                    [["C1", "A", "", "F1", "2026-07-01", "2026-07-31", "TRY", "10", "0", "", ""]],
                ),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        format="multipart",
    )
    assert bad_name.status_code == status.HTTP_400_BAD_REQUEST
    assert bad_name.data["code"] == "dangerous_filename"


@pytest.mark.django_db
def test_row_validation_messages(org_owner):
    """NP-064 Turkish validation messages."""
    org, _owner = org_owner
    base = {
        "müşteri_kodu": "C-1",
        "müşteri_adı": "Cari",
        "fatura_numarası": "F-1",
        "fatura_tarihi": "2026-07-01",
        "vade_tarihi": "2026-07-31",
        "fatura_tutarı": "100",
        "para_birimi": "TRY",
        "ödenen_tutar": "0",
        "vergi_numarası": "",
        "telefon": "",
        "email": "",
    }

    empty_number = validate_mapped_row(
        row_number=2,
        mapped={**base, "fatura_numarası": ""},
        organization=org,
    )
    assert any(i.error_message == "Boş fatura numarası" for i in empty_number.issues)

    bad_date = validate_mapped_row(
        row_number=2,
        mapped={**base, "fatura_tarihi": "not-a-date"},
        organization=org,
    )
    assert any(i.error_message == "Geçersiz tarih" for i in bad_date.issues)

    bad_amount = validate_mapped_row(
        row_number=2,
        mapped={**base, "fatura_tutarı": "abc"},
        organization=org,
    )
    assert any(i.error_message == "Tutar sayısal değil" for i in bad_amount.issues)

    negative = validate_mapped_row(
        row_number=2,
        mapped={**base, "fatura_tutarı": "-10"},
        organization=org,
    )
    assert any(i.error_message == "Negatif tutar" for i in negative.issues)

    due_before = validate_mapped_row(
        row_number=2,
        mapped={**base, "vade_tarihi": "2026-06-01"},
        organization=org,
    )
    assert any(
        i.error_message == "Vade tarihi fatura tarihinden önce" for i in due_before.issues
    )

    missing_customer = validate_mapped_row(
        row_number=2,
        mapped={**base, "müşteri_kodu": "MISSING", "müşteri_adı": ""},
        organization=org,
    )
    assert any(i.error_message == "Müşteri bulunamadı" for i in missing_customer.issues)

    customer = Customer.objects.create(organization=org, code="C-1", name="Cari")
    Invoice.objects.create(
        organization=org,
        customer=customer,
        number="F-1",
        invoice_date=date(2026, 7, 1),
        due_date=date(2026, 7, 31),
        total_amount=Decimal("100.00"),
        status=InvoiceStatus.OPEN,
    )
    dup = validate_mapped_row(row_number=2, mapped=base, organization=org)
    assert dup.is_duplicate
    assert any(i.error_message == "Aynı fatura daha önce eklenmiş" for i in dup.issues)


@pytest.mark.django_db
def test_upload_mapping_preview_commit_flow(api_client, org_owner, settings):
    settings.CELERY_TASK_ALWAYS_EAGER = True
    org, owner = org_owner
    client = _auth(api_client, owner, org)

    headers = ["Cari Ünvanı", "Belge No", "Vade", "Borç", "Fatura Tarihi", "Kod"]
    rows = [
        ["Yeni Cari A.Ş.", "F-100", "2026-08-15", "2500,50", "2026-07-01", "YC-1"],
        # NP-065 within-file duplicate (same key)
        ["Yeni Cari A.Ş.", "F-100", "2026-08-15", "2500,50", "2026-07-01", "YC-1"],
        # Invalid amount
        ["Yeni Cari A.Ş.", "F-101", "2026-08-15", "abc", "2026-07-01", "YC-1"],
    ]
    content = _xlsx_bytes(headers, rows)
    upload = client.post(
        "/api/imports/invoices/upload/",
        {
            "file": SimpleUploadedFile(
                "cari.xlsx",
                content,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        format="multipart",
    )
    assert upload.status_code == status.HTTP_201_CREATED, upload.data
    job_id = upload.data["job"]["id"]
    assert upload.data["job"]["status"] == ImportJobStatus.PENDING
    assert upload.data["job"]["total_rows"] == 3

    dup = client.post(
        "/api/imports/invoices/upload/",
        {
            "file": SimpleUploadedFile(
                "cari2.xlsx",
                content,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        format="multipart",
    )
    assert dup.status_code == status.HTTP_400_BAD_REQUEST
    assert dup.data["code"] == "duplicate_file"

    mapping = _full_mapping(
        {
            "müşteri_adı": "Cari Ünvanı",
            "müşteri_kodu": "Kod",
            "fatura_numarası": "Belge No",
            "vade_tarihi": "Vade",
            "fatura_tutarı": "Borç",
            "fatura_tarihi": "Fatura Tarihi",
        }
    )
    mapped = client.patch(
        f"/api/imports/{job_id}/mapping/",
        {"column_mapping": mapping},
        format="json",
    )
    assert mapped.status_code == status.HTTP_200_OK
    assert mapped.data["job"]["status"] == ImportJobStatus.PENDING

    preview = client.post(
        f"/api/imports/{job_id}/preview/",
        {"duplicate_policy": DuplicatePolicy.SKIP},
        format="json",
    )
    assert preview.status_code == status.HTTP_200_OK, preview.data
    summary = preview.data["summary"]
    assert summary["total_rows"] == 3
    assert summary["invalid_rows"] == 1
    assert summary["likely_duplicate_count"] >= 1
    assert preview.data["job"]["status"] == ImportJobStatus.READY
    assert Customer.objects.filter(organization=org).count() == 0
    assert Invoice.objects.filter(organization=org).count() == 0

    commit = client.post(
        f"/api/imports/{job_id}/commit/",
        {"duplicate_policy": DuplicatePolicy.SKIP},
        format="json",
    )
    assert commit.status_code == status.HTTP_202_ACCEPTED, commit.data

    detail = client.get(f"/api/imports/{job_id}/")
    assert detail.status_code == status.HTTP_200_OK
    assert detail.data["status"] == ImportJobStatus.COMPLETED
    assert detail.data["successful_rows"] == 1
    assert detail.data["skipped_duplicates"] >= 1
    assert detail.data["failed_rows"] == 1
    assert Invoice.objects.filter(organization=org).count() == 1
    assert Customer.objects.filter(organization=org).count() == 1

    export = client.get(f"/api/imports/{job_id}/errors/export/")
    assert export.status_code == status.HTTP_200_OK
    assert export.content[:2] == b"PK"
    wb = load_workbook(io.BytesIO(export.content))
    assert wb.active.max_row >= 2


def test_suggest_mapping_aliases():
    mapping = suggest_mapping(["Cari Ünvanı", "Belge No", "Vade", "Borç"])
    assert mapping["müşteri_adı"] == "Cari Ünvanı"
    assert mapping["fatura_numarası"] == "Belge No"
    assert mapping["vade_tarihi"] == "Vade"
    assert mapping["fatura_tutarı"] == "Borç"
